from typing import Any
from dataclasses import dataclass

import numpy as np
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Border, Side
from openpyxl.worksheet.worksheet import Worksheet

from deicide.validation2 import Clustering

@dataclass
class ContiguousRange:
    start: int
    end: int
    value: Any

    def __init__(self, start: int, end: int, value: Any):
        if start >= end:
            raise ValueError("invalid range")
        self.start = start
        self.end = end
        self.value = value

    def offset(self, delta: int) -> "ContiguousRange":
        return ContiguousRange(self.start + delta, self.end + delta, self.value)

    def is_zero(self) -> bool:
        return self.value == 0

    def is_unit(self) -> bool:
        return len(self) == 1

    def __len__(self) -> int:
        return self.end - self.start


@dataclass
class MergeRect:
    start_i: int
    start_j: int
    end_i: int
    end_j: int

    def __init__(self, start_i: int, start_j: int, end_i: int, end_j: int):
        self.start_i = start_i
        self.start_j = start_j
        self.end_i = end_i
        self.end_j = end_j

    def offset(self, delta_i: int, delta_j: int) -> "MergeRect":
        return MergeRect(
            self.start_i + delta_i,
            self.start_j + delta_j,
            self.end_i + delta_i,
            self.end_j + delta_j,
        )


def find_contiguous(arr: list[Any]) -> list[ContiguousRange]:
    ret = []
    if len(arr) == 0:
        return ret
    init_index, init_value = 0, arr[0]
    for curr_index in range(1, len(arr)):
        curr_value = arr[curr_index]
        if curr_value != init_value:
            ret.append(ContiguousRange(init_index, curr_index, init_value))
            init_index, init_value = curr_index, curr_value
    ret.append(ContiguousRange(init_index, len(arr), init_value))
    return ret


def find_row_merges(mat: np.ndarray) -> list[MergeRect]:
    n_rows, _ = mat.shape
    if n_rows == 0:
        return []
    ranges = find_contiguous(list(mat[0]))
    ranges = [r for r in ranges if not r.is_unit()]
    merges = [MergeRect(0, r.start, 1, r.end) for r in ranges]
    if n_rows == 1:
        return merges
    for r in ranges:
        indices = list(range(r.start, r.end))
        merges += [m.offset(1, r.start) for m in find_row_merges(mat[1:, indices])]
    return merges


def find_merges(mat: np.ndarray) -> list[MergeRect]:
    _, n_cols = mat.shape
    binary_mat = np.ones(shape=mat.shape)
    merges = find_row_merges(mat)

    for merge in merges:
        for i in range(merge.start_i, merge.end_i):
            for j in range(merge.start_j, merge.end_j):
                binary_mat[i, j] = 0

    for j in range(n_cols):
        ranges = find_contiguous(list(binary_mat[:, j]))
        ranges = list(filter(lambda r: not (r.is_unit() or r.is_zero()), ranges))
        merges += [MergeRect(r.start, j, r.end, j + 1) for r in ranges]

    return merges


def add_header_cells(ws: Worksheet, mat: np.ndarray, delta_i: int, delta_j: int):
    n_rows, n_cols = mat.shape
    thin = Side(border_style="thin", color="000000")
    for i in range(n_rows):
        for j in range(n_cols):
            value = mat[i][j]
            cell = ws.cell(row=i + delta_i + 1, column=j + delta_j + 1, value=value)
            text_rotation = 180 if len(str(value)) > 4 else 0
            cell.alignment = Alignment(
                horizontal="center", vertical="top", text_rotation=text_rotation
            )
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)


def merge_header_cells(ws: Worksheet, merges: list[MergeRect]):
    for rect in merges:
        ws.merge_cells(
            start_row=rect.start_i + 1,
            start_column=rect.start_j + 1,
            end_row=rect.end_i,
            end_column=rect.end_j,
        )


def add_clustering(ws: Worksheet, clustering: Clustering, sort_entities, get_name):
    filled_clustering = clustering.fill("#")
    cluster_names = sorted(c.name for c in filled_clustering.clusters())
    mat = np.atleast_2d(np.array([c.split("$") for c in cluster_names])).T
    merges = [m.offset(0, 1) for m in find_merges(mat)]
    add_header_cells(ws, mat, delta_i=0, delta_j=1)
    merge_header_cells(ws, merges)
    delta_i = mat.shape[0]
    entity_ids = sort_entities(filled_clustering.entities())
    for i, entity_id in enumerate(entity_ids):
        ws.cell(delta_i + i + 1, 1, get_name(entity_id))
    for e, c in filled_clustering.pairs:
        ws.cell(delta_i + entity_ids.index(e) + 1, cluster_names.index(c.name) + 2, "T")


def dump_indicators(entities_df, path, **kwargs):
    def sort_entities(entitiy_ids: set[int]) -> list[int]:
        return sorted(entitiy_ids, key=lambda id: entities_df.loc[id]["start_row"])

    def get_name(entity_id: int) -> str:
        return entities_df.loc[entity_id]["name"]

    wb = Workbook()
    for name, clustering in kwargs.items():
        wb.create_sheet(name)
        add_clustering(wb[name], clustering, sort_entities, get_name)

    del wb["Sheet"]
    wb.save(path)
